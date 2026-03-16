#!/usr/bin/env python3
"""
Generate an Excel template file from the CSV templates.
This creates a multi-sheet Excel file ready to upload to Google Drive.

Requirements:
    pip install openpyxl

Usage:
    python generate_excel_template.py

Output:
    Creates "Expense_Tracker_Template.xlsx" with all three sheets
"""

import csv
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl not installed. Install with: pip install openpyxl")
    exit(1)


def read_csv(filepath):
    """Read CSV file and return rows"""
    with open(filepath, 'r') as f:
        reader = csv.reader(f)
        return list(reader)


def create_excel_template():
    """Create Excel template from CSV files"""
    script_dir = Path(__file__).parent
    output_file = script_dir / "Expense_Tracker_Template.xlsx"

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Define sheets to create
    sheets = [
        ("Subscriptions", "Subscriptions.csv", 16),  # 16 columns (A-P)
        ("One-Time Purchases", "One-Time_Purchases.csv", 10),  # 10 columns (A-J)
        ("Logs", "Logs.csv", 5),  # 5 columns (A-E)
    ]

    for sheet_name, csv_file, num_cols in sheets:
        print(f"Creating '{sheet_name}' sheet...")

        # Read CSV data
        csv_path = script_dir / csv_file
        if not csv_path.exists():
            print(f"⚠️  Warning: {csv_file} not found, skipping...")
            continue

        rows = read_csv(csv_path)

        # Create sheet
        ws = wb.create_sheet(sheet_name)

        # Write all rows
        for row_idx, row_data in enumerate(rows, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Check if it's a formula (starts with =)
                if value.startswith('='):
                    cell.value = value
                else:
                    cell.value = value

                # Style header row
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

        # Auto-adjust column widths
        for col_idx in range(1, num_cols + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0

            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

    # Save workbook
    wb.save(output_file)
    print(f"\n✅ Excel template created: {output_file}")
    print(f"\nNext steps:")
    print(f"1. Upload '{output_file.name}' to Google Drive")
    print(f"2. Open in Google Sheets")
    print(f"3. Copy the spreadsheet ID from the URL")
    print(f"4. Update your .env file with GOOGLE_SPREADSHEET_ID")


if __name__ == "__main__":
    create_excel_template()
